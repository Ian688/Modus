"""office_exec: LLM-reasoned, sandboxed Office scripting (Phase A3.2).

Gives the LLM a base for arbitrary Office operations (aggregate/filter/format)
instead of one tool per scenario.  The script runs against ONE PathGuard-validated
file; dangerous imports are AST-rejected; write-looking scripts are
approval-gated.
"""

from __future__ import annotations

import pytest

from modus.config import ModusConfig
from modus.policy.approval import ApprovalDecision, ApprovalPolicy
from modus.tools.base import ToolContext
from modus.tools.builtins import get_builtin_tools
from modus.tools.office_exec import (
    _MAX_SCRIPT_CHARS,
    _reject_dangerous_imports,
    _run_script,
    office_exec,
)


@pytest.fixture
def ows(tmp_path, monkeypatch):
    from pathlib import Path

    monkeypatch.setattr(Path, "home", classmethod(lambda cls: tmp_path))
    ws = tmp_path / "ws"
    ws.mkdir()
    return ws


def _ctx(ws):
    return ToolContext(cwd=str(ws), workspace_root=str(ws), config=ModusConfig())


def _make_excel(ws):
    import openpyxl

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.append(["region", "amount"])
    for i in range(100):
        sh.append(["east" if i % 2 else "west", i * 10])
    p = ws / "sales.xlsx"
    wb.save(p)
    return p


# ── declaration contract ──


def test_office_exec_declared():
    tools = {t.name: t for t in get_builtin_tools()}
    assert "office_exec" in tools
    tool = tools["office_exec"]
    assert tool.is_read_only is False
    assert tool.danger_level == "medium"
    assert tool.requires_approval is True
    assert "exec" in tool.capabilities
    assert "filesystem" in tool.capabilities


def test_office_exec_approval_policy():
    tools = {t.name: t for t in get_builtin_tools()}
    policy = ApprovalPolicy(ModusConfig().policy)
    assert policy.evaluate(tools["office_exec"]) is ApprovalDecision.ASK


# ── import rejection (AST scan) ──


def test_reject_dangerous_imports():
    assert _reject_dangerous_imports("import subprocess\nprint('x')") is not None
    assert _reject_dangerous_imports("from socket import socket") is not None
    assert _reject_dangerous_imports("import urllib.request") is not None
    assert _reject_dangerous_imports("import os\nprint('ok')") is None
    assert _reject_dangerous_imports("import openpyxl\nprint('ok')") is None
    assert _reject_dangerous_imports("from collections import defaultdict") is None


def test_reject_dangerous_imports_syntax_error():
    assert _reject_dangerous_imports("def broken(") is not None


# ── execution against a real file ──


@pytest.mark.asyncio
async def test_office_exec_aggregates_excel(ows):
    _make_excel(ows)
    result = await office_exec({"path": "sales.xlsx", "script": '''
import openpyxl, json
from collections import defaultdict
wb = openpyxl.load_workbook(ABS_PATH, read_only=True, data_only=True)
ws = wb.active
totals = defaultdict(int)
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]: totals[row[0]] += row[1] or 0
print(json.dumps(dict(totals)))
'''}, _ctx(ows))
    assert not result.is_error
    assert "east" in result.content
    assert "west" in result.content


@pytest.mark.asyncio
async def test_office_exec_rejects_blocked_import(ows):
    _make_excel(ows)
    result = await office_exec(
        {"path": "sales.xlsx", "script": "import subprocess\nprint('x')"}, _ctx(ows),
    )
    assert result.is_error
    assert "not allowed" in result.content


@pytest.mark.asyncio
async def test_office_exec_script_too_long(ows):
    _make_excel(ows)
    result = await office_exec(
        {"path": "sales.xlsx", "script": "x" * (_MAX_SCRIPT_CHARS + 1)}, _ctx(ows),
    )
    assert result.is_error
    assert "exceeds" in result.content


@pytest.mark.asyncio
async def test_office_exec_missing_file(ows):
    result = await office_exec({"path": "nope.xlsx", "script": "print('x')"}, _ctx(ows))
    assert result.is_error


@pytest.mark.asyncio
async def test_office_exec_unsupported_format(ows):
    (ows / "data.txt").write_text("x", encoding="utf-8")
    result = await office_exec({"path": "data.txt", "script": "print('x')"}, _ctx(ows))
    assert result.is_error
    assert "unsupported" in result.content


def test_run_script_uses_venv_interpreter(ows):
    """The sandbox runs under Modus's venv so openpyxl is importable."""
    _make_excel(ows)
    target = ows / "sales.xlsx"
    out = _run_script(target, "print('venv-ok')")
    assert "venv-ok" in out.stdout


# ── sandbox hardening (Phase A3.2 deep-dive) ──


def test_reject_os_file_operations():
    assert _reject_dangerous_imports("os.remove('x')") is not None
    assert _reject_dangerous_imports("import os\nos.unlink('/etc/passwd')") is not None
    assert _reject_dangerous_imports("os.system('rm -rf /')") is not None
    assert _reject_dangerous_imports("shutil.rmtree('/tmp/x')") is not None


def test_reject_os_path_operations_allowed():
    # Path operations on the target are fine (not in the blocklist).
    assert _reject_dangerous_imports("import os\nos.path.exists(PATH)") is None
    assert _reject_dangerous_imports("import os\nos.path.basename(ABS_PATH)") is None


def test_reject_shutil_attr():
    assert _reject_dangerous_imports("import shutil\nshutil.copy('a','b')") is not None


@pytest.mark.asyncio
async def test_office_exec_timeout_kills_process_group(ows, monkeypatch):
    """A hanging script is terminated at the timeout, not left running."""
    import modus.tools.office_exec as oe

    _make_excel(ows)
    # Speed up the timeout for the test.
    monkeypatch.setattr(oe, "_TIMEOUT", 1.0)
    result = await office_exec(
        {"path": "sales.xlsx", "script": "import time\ntime.sleep(30)\nprint('never')"}, _ctx(ows),
    )
    assert result.is_error
    # SIGKILL (-9) from the process-group kill, surfaced as a failed run.
    assert "exit code -9" in result.content
